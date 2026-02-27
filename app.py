import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import io
import os

def generate_excel(input_file):
    # Διάβασμα του υπάρχοντος Excel
    df = pd.read_excel(input_file, header=None)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Υπολογισμός Μισθοδοσίας"

    # --- ΟΡΙΣΜΟΣ ΣΤΥΛ ---
    thin_side = Side(border_style="thin", color="000000")
    border_all = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    italic_font = Font(italic=True, size=10, color="444444")
    red_font = Font(color="FF0000", bold=True)
    
    currency_format = '#,##0.00€'
    integer_format = '0'

    # Πλάτος στηλών
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 45

    # --- ΕΠΕΞΕΡΓΑΣΙΑ 242 ΓΡΑΜΜΩΝ ---
    # Χρησιμοποιούμε το index του dataframe για να γεμίσουμε το Excel
    for index, row in df.iterrows():
        excel_row = index + 1
        if excel_row > 242:
            break
        
        # Στήλη Β: Περιγραφή (Ακριβής μεταφορά από Column 1 του df)
        desc = row[1] if len(row) > 1 and pd.notnull(row[1]) else ""
        ws.cell(row=excel_row, column=2).value = desc
        
        # Στήλη D: Τιμές (Column 3 του df)
        val_d = row[3] if len(row) > 3 and pd.notnull(row[3]) else ""
        if val_d != "":
            ws.cell(row=excel_row, column=4).value = val_d
            # Μορφοποίηση αριθμών
            if isinstance(val_d, (int, float)) and val_d != 0:
                if abs(val_d) > 163 or not float(val_d).is_integer():
                    ws.cell(row=excel_row, column=4).number_format = currency_format
                else:
                    ws.cell(row=excel_row, column=4).number_format = integer_format

        # Στήλες F & G: Σημειώσεις (Italics) - Columns 5 και 6 του df
        for col_idx, df_col in zip([6, 7], [5, 6]):
            if len(row) > df_col:
                col_val = row[df_col]
                if pd.notnull(col_val):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.value = str(col_val)
                    cell.font = italic_font
                    cell.alignment = Alignment(wrap_text=True)

        # Εφαρμογή πλαισίου σε κάθε γραμμή ξεχωριστά (A-G)
        for col in range(1, 8):
            ws.cell(row=excel_row, column=col).border = border_all

    # --- ΕΞΕΙΔΙΚΕΥΜΕΝΑ DROPDOWNS ---
    # D3: Μισθολογικό Κλιμάκιο
    dv_d3 = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10,11,12,13,14,15"', allow_blank=True)
    ws.add_data_validation(dv_d3)
    dv_d3.add(ws['D3'])

    # D5: Επίδομα Γάμου
    dv_d5 = DataValidation(type="list", formula1='"ΝΑΙ,ΟΧΙ"', allow_blank=True)
    ws.add_data_validation(dv_d5)
    dv_d5.add(ws['D5'])

    # D7: Πολυετία
    dv_d7 = DataValidation(type="list", formula1='"0,5,10,15,20,25,30"', allow_blank=True)
    ws.add_data_validation(dv_d7)
    dv_d7.add(ws['D7'])

    # D20: Τύπος Εργασίας
    dv_d20 = DataValidation(type="list", formula1='"ΠΛΗΡΕΣ,ΜΕΙΩΜΕΝΟ,ΕΚΤΑΚΤΟ"', allow_blank=True)
    ws.add_data_validation(dv_d20)
    dv_d20.add(ws['D20'])

    # --- ΕΛΕΓΧΟΣ ΩΡΩΝ D15-D17 ---
    ws['E17'].formula = '=IF(SUM(D15:D17)<>162.5, "⚠ ΑΘΡΟΙΣΜΑ ΩΡΩΝ ≠ 162.50", "")'
    ws['E17'].font = red_font

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- STREAMLIT INTERFACE ---
st.set_page_config(page_title="Georgilas Salary App", layout="wide")
st.title("📊 Salary Calculator Excel Generator")

# Το όνομα του αρχείου όπως είναι στο Repo σου
FILE_NAME = 'Georgilas Salary Calc.xlsx'

if os.path.exists(FILE_NAME):
    st.success(f"Το αρχείο '{FILE_NAME}' βρέθηκε στο αποθετήριο.")
    if st.button('🚀 Δημιουργία Τελικού Excel'):
        try:
            excel_data = generate_excel(FILE_NAME)
            st.download_button(
                label="📥 Κατέβασμα Διορθωμένου Αρχείου",
                data=excel_data,
                file_name="Georgilas_Salary_Final_Formatted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Σφάλμα: {e}")
else:
    st.error(f"Το αρχείο '{FILE_NAME}' δεν βρέθηκε στο φάκελο. Βεβαιωθείτε ότι το όνομα στο GitHub είναι ακριβώς το ίδιο.")
    st.info("Εναλλακτικά, ανεβάστε το χειροκίνητα εδώ:")
    manual_file = st.file_uploader("Upload Excel", type=["xlsx"])
    if manual_file and st.button("Generate from Upload"):
        excel_data = generate_excel(manual_file)
        st.download_button("Download", excel_data, "Salary_Calc.xlsx")
